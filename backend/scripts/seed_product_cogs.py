"""Seed the product_cogs table from `COGS Magical Butter.xlsx`.

Schema discovered from the workbook:
  - US / CA / UK sheets: Product Name | ASIN | SKU | Status | Normal Price | COGS
  - AU sheet:            Product Name | ASIN | SKU | Normal Price | COGS  (no Status)

A row is "active" when COGS is a positive number AND (Status is not provided
OR Status is non-empty and not literally "Discontinued"). Discontinued items
have no usable COGS in this seller's process and are skipped per
PNL_MAPPING.md.

Idempotency: deletes existing rows for each marketplace before inserting,
so re-running with an updated spreadsheet replaces (not duplicates) the
prior seed.

Run from backend/ with .env loaded:
    .venv/bin/python -m scripts.seed_product_cogs
"""

from __future__ import annotations

import asyncio
import logging
import sys
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

import openpyxl
from sqlalchemy import delete, select, func

from app.database import AsyncSessionLocal
from app.models import ProductCogs

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

WORKBOOK_PATH = Path("/Volumes/ACERFD/TECH/MYAMZTEAM/myamzteam-erp/COGS Magical Butter.xlsx")

SHEET_CONFIG: dict[str, dict[str, str]] = {
    "US": {"marketplace": "us", "currency": "USD"},
    "CA": {"marketplace": "ca", "currency": "CAD"},
    "UK": {"marketplace": "uk", "currency": "GBP"},
    "AU": {"marketplace": "au", "currency": "AUD"},
}


def _to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _to_sku(value) -> str | None:
    if value is None:
        return None
    # SKUs read as float (e.g. 850251005008.0) need integer-string normalization
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = str(value).strip()
    return s or None


def _is_active(status_cell, cogs: Decimal | None) -> bool:
    if cogs is None or cogs <= 0:
        return False
    if status_cell is None:
        return True
    status = str(status_cell).strip().lower()
    if not status:
        return True
    # Most CA rows put the product name in the Status column rather than a
    # status keyword; treat that as active. Only explicit discontinued
    # keywords disqualify a row.
    return "discontinued" not in status and "inactive" not in status


def _iter_sheet_rows(ws) -> Iterable[dict]:
    """Yield {product_name, asin, sku, status_cell, price, cogs} dicts."""
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    # Build a header-name -> index map so we tolerate the AU sheet having no Status column.
    idx = {h.lower(): i for i, h in enumerate(headers)}

    def get(row, key_substr: str, default=None):
        for k, i in idx.items():
            if key_substr in k:
                return row[i]
        return default

    for row in ws.iter_rows(min_row=2, values_only=True):
        yield {
            "product_name": get(row, "product name"),
            "asin": get(row, "asin"),
            "sku": get(row, "sku"),
            "status_cell": get(row, "status"),  # None for AU
            "price": get(row, "normal product price") or get(row, "price"),
            "cogs": get(row, "cogs"),
        }


async def main() -> int:
    if not WORKBOOK_PATH.exists():
        print(f"ERROR: spreadsheet not found at {WORKBOOK_PATH}")
        return 1

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    today = date.today()

    counts: dict[str, dict[str, int]] = {}
    async with AsyncSessionLocal() as session:
        for sheet_name, cfg in SHEET_CONFIG.items():
            if sheet_name not in wb.sheetnames:
                logger.warning("seed_product_cogs sheet_missing=%s", sheet_name)
                continue
            ws = wb[sheet_name]
            marketplace = cfg["marketplace"]
            currency = cfg["currency"]
            inserted = 0
            skipped_inactive = 0
            skipped_missing = 0
            seen_keys: set[str] = set()

            # Purge prior rows for this marketplace so re-runs replace (idempotency).
            await session.execute(delete(ProductCogs).where(ProductCogs.marketplace == marketplace))

            for row in _iter_sheet_rows(ws):
                sku = _to_sku(row["sku"])
                cogs = _to_decimal(row["cogs"])
                price = _to_decimal(row["price"])
                asin = _to_sku(row["asin"])
                product_name = (str(row["product_name"]).strip() if row["product_name"] else None)

                if not sku:
                    skipped_missing += 1
                    continue
                if not _is_active(row["status_cell"], cogs):
                    skipped_inactive += 1
                    continue
                # De-dupe within a sheet -- the workbook has occasional duplicate SKU rows
                if sku in seen_keys:
                    continue
                seen_keys.add(sku)

                session.add(
                    ProductCogs(
                        marketplace=marketplace,
                        sku=sku,
                        asin=asin,
                        product_name=product_name,
                        unit_cost=cogs,
                        product_price=price,
                        currency=currency,
                        status="active",
                        effective_date=today,
                    )
                )
                inserted += 1

            counts[sheet_name] = {
                "inserted": inserted,
                "skipped_inactive_or_zero_cogs": skipped_inactive,
                "skipped_no_sku": skipped_missing,
            }

        await session.commit()

    # Print summary
    print()
    print(f"{'sheet':6}  {'inserted':>9}  {'skipped_inactive':>18}  {'skipped_no_sku':>15}")
    for sheet, c in counts.items():
        print(f"{sheet:6}  {c['inserted']:>9}  {c['skipped_inactive_or_zero_cogs']:>18}  {c['skipped_no_sku']:>15}")

    # Verify table state
    async with AsyncSessionLocal() as session:
        s = select(ProductCogs.marketplace, func.count(), func.coalesce(func.avg(ProductCogs.unit_cost), 0)).group_by(ProductCogs.marketplace)
        print()
        print(f"{'mkt':4}  {'rows':>6}  {'avg_cogs':>10}")
        for mp, cnt, avg in (await session.execute(s)).all():
            print(f"{mp:4}  {cnt:>6}  {float(avg):>10,.2f}")

    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
